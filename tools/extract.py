# -*- coding: utf-8 -*-
"""Извлечение данных по линиям: TESA REGISTER v65 + ЧЕМОДАНЫ.xlsx -> seed JSON"""
import openpyxl, re, json, math, collections, sys, os

REG = '/mnt/user-data/uploads/Tervasaari Pr/TESA REGISTER - полный v65.xlsx'
BOXES = '/mnt/user-data/uploads/Tervasaari Pr/ЧЕМОДАНЫ.xlsx'
OUT = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), 'data', 'seed.json')
DRAWINGS = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), 'public', 'drawings')

# ---------- 0. чертежи (PDF изометрий) ----------
draw = {}
if os.path.isdir(DRAWINGS):
    for fn in sorted(os.listdir(DRAWINGS)):
        if not fn.lower().endswith('.pdf'):
            continue
        m0 = re.match(r'^(50\d{3})', fn)
        if m0:
            rev = re.search(r'_R(\d+)', fn)
            draw[m0.group(1)] = {'file': fn, 'rev': ('R' + rev.group(1)) if rev else ''}
print('чертежей найдено:', len(draw))

# ---------- 1. чемоданы (съёмные короба на арматуру) из ЧЕМОДАНЫ.xlsx ----------
# «Общий Чистый» — сводный список: линия, кол-во, диаметр, толщина, номер, примечание
# «Основные»     — список арматуры Valmet: позиция, тип, изготовитель
wbb = openpyxl.load_workbook(BOXES, read_only=True, data_only=True)

def _s(v):
    return '' if v is None else str(v).strip()

def _i(v, d=0):
    try:
        return int(round(float(v)))
    except (TypeError, ValueError):
        return d

# арматура Valmet со съёмной изоляцией: по линии — позиции
valves_by_line = collections.defaultdict(list)
for r in wbb['Основные '].iter_rows(min_row=4, values_only=True):
    if not r[6] or 'emovable' not in _s(r[17]):
        continue
    m = re.match(r'^(50\d{3})', _s(r[6]))
    if not m:
        continue
    valves_by_line[m.group(1)].append({
        'pos': _s(r[3]), 'size': _s(r[9]), 'conn': _s(r[10]),
        'valve': _s(r[11]), 'note': _s(r[15]),
    })

boxes_by_line = collections.defaultdict(list)
kot = collections.Counter()          # число чемоданов на линию
for r in wbb['Общий Чистый'].iter_rows(min_row=3, values_only=True):
    if not r[1]:
        continue
    m = re.match(r'^(50\d{3})', _s(r[1]))
    if not m:
        continue
    num = m.group(1)
    qty = _i(r[4], 0)
    if qty <= 0:
        continue
    box = {
        'n': _s(r[7]),                # BOX - ###  (номер по проекту)
        'qty': qty,
        'dn': _s(r[5]),               # BOX - pipe (диаметр арматуры)
        'ins': _i(r[6], 0),           # BOX - EV   (толщина изоляции, мм)
        'info': _s(r[8]),             # тип/примечание
        'measured': bool(_s(r[9])),
        'ordered': bool(_s(r[10])),
        'installed': bool(_s(r[11])),
    }
    # подтягиваем позицию арматуры из списка Valmet по совпадению диаметра
    for v in valves_by_line.get(num, []):
        if v.get('used'):
            continue
        if v['size'] and box['dn'] and v['size'].split('/')[0] == box['dn'].split('/')[0]:
            v['used'] = True
            box['pos'] = v['pos']
            box['valve'] = v['valve'] or v['note']
            box['conn'] = v['conn']
            break
    boxes_by_line[num].append(box)
    kot[num] += qty

print('чемоданов:', sum(kot.values()), 'на линиях:', len(kot))
wbb.close()

# ---------- 2. регистр ----------
wb2 = openpyxl.load_workbook(REG, data_only=True)
ws = wb2['REGISTERFULL']

C = dict(line=0, otemp=1, dtemp=2, part=3, dn=4, heat=5, typ=6, id=7, ins=8, od=9,
         dev=10, qty=11, r=12, deg=13, lmm=14, summ=15, cone=16, al=17, partial=18,
         err=19, connect=20, release=21, relpct=22)

def f(v, d=0.0):
    try:
        return float(v)
    except (TypeError, ValueError):
        return d

OD2DN = {21: 15, 22: 15, 26: 20, 27: 20, 33: 25, 34: 25, 42: 32, 48: 40,
         60: 50, 76: 65, 89: 80, 114: 100, 140: 125, 139: 125, 168: 150,
         219: 200, 273: 250, 323: 300, 324: 300, 355: 350, 356: 350, 406: 400}

def dn_of(od, fallback=None):
    if not od:
        return fallback
    o = int(round(od))
    if o in OD2DN:
        return OD2DN[o]
    best, bd = fallback, 1e9
    for k, v in OD2DN.items():
        if abs(k - o) < bd:
            bd, best = abs(k - o), v
    return best if bd <= 4 else fallback

ELBOW = {'Elbow'}
PIPE = {'Pipe'}
TIE = {'TEE', 'Врезка', 'Set on Tee'}
CONE = {'Cone Flat', 'Cone Conc.'}
VALVE = {'VALVE'}
FLANGE = {'Flange WN'}

lines = collections.OrderedDict()
for r in ws.iter_rows(min_row=3, values_only=True):
    name = r[C['line']]
    if not name:
        continue
    name = str(name).strip()
    L = lines.setdefault(name, {
        'name': name, 'rows': [], 'connect': set(), 'al': set(),
        'otemp': None, 'dtemp': None, 'partial': False,
        'rel_dates': set(), 'rel_notes': set(), 'rel_pct': []})
    part = str(r[C['part']] or '').strip()
    qty = int(f(r[C['qty']], 0) or 0)
    summ = f(r[C['summ']])
    dev = f(r[C['dev']])
    row = {
        'part': part, 'dn': r[C['dn']], 'id': f(r[C['id']]), 'ins': f(r[C['ins']]),
        'od': f(r[C['od']]), 'dev': dev, 'qty': qty, 'deg': r[C['deg']],
        'lmm': f(r[C['lmm']]), 'summ': summ, 'cone': r[C['cone']],
        'al': (str(r[C['al']]).strip() if r[C['al']] else ''),
        'area': round(summ * dev / 1000.0, 3),
    }
    L['rows'].append(row)
    if row['al']:
        L['al'].add(re.sub(r'\s*мм', '', row['al']).strip())
    if r[C['otemp']] and not L['otemp']:
        L['otemp'] = str(r[C['otemp']]).strip()
    if r[C['dtemp']] and not L['dtemp']:
        L['dtemp'] = str(r[C['dtemp']]).strip()
    p = str(r[C['partial']] or '').strip().lower()
    if p not in ('no', 'нет', ''):
        L['partial'] = True
    rel = r[C['release']]
    if rel not in (None, ''):
        t = str(rel).strip()
        if t:
            m = re.match(r'^(\d{4})-(\d{2})-(\d{2})', t)
            if m:
                L['rel_dates'].add(m.group(0))
            else:
                L['rel_notes'].add(t)
    rp = r[C['relpct']]
    if rp not in (None, ''):
        try:
            L['rel_pct'].append(float(rp))
        except (TypeError, ValueError):
            pass

    conn = r[C['connect']]
    if conn:
        for tok in re.split(r'[,;/]+', str(conn)):
            tok = tok.strip()
            if tok:
                L['connect'].add(tok)

out = []
for name, L in lines.items():
    num = re.match(r'^(50\d{3})', name)
    num = num.group(1) if num else name[:5]
    rows = L['rows']
    length_m = round(sum(x['summ'] for x in rows), 2)
    area_m2 = round(sum(x['area'] for x in rows), 2)
    straight_m = round(sum(x['summ'] for x in rows if x['part'] in PIPE), 2)
    elbows = sum(x['qty'] for x in rows if x['part'] in ELBOW)
    ties = sum(x['qty'] for x in rows if x['part'] in TIE)
    cones = sum(x['qty'] for x in rows if x['part'] in CONE)
    valves_reg = sum(x['qty'] for x in rows if x['part'] in VALVE)
    flanges = sum(x['qty'] for x in rows if x['part'] in FLANGE)
    cases = kot.get(num, 0)          # съёмные чемоданы по списку ЧЕМОДАНЫ.xlsx

    # --- расчётная площадь для весов (не искажает отчётные м²) ---
    def work_area(x):
        a = x['area']
        if a > 0:
            return a
        dev = x['dev'] / 1000.0
        if dev <= 0:
            return 0.0
        q = max(x['qty'], 1)
        if x['part'] in ELBOW:
            deg = 90.0
            try:
                deg = float(x['deg'])
            except (TypeError, ValueError):
                deg = 90.0
            R = 1.5 * (x['id'] or 0)
            L = math.pi * R * deg / 180.0 / 1000.0
            return round(L * dev * q, 3)
        if x['part'] in VALVE:
            return round(0.8 * dev * q, 3)
        if x['part'] in FLANGE:
            return round(0.35 * dev * q, 3)
        if x['part'] in TIE | CONE:
            return round(0.3 * dev * q, 3)
        return 0.0

    for x in rows:
        x['work'] = work_area(x)

    # площади по группам (для весов внутри подпозиций)
    a_elbow = round(sum(x['area'] for x in rows if x['part'] in ELBOW), 2)
    a_pipe = round(sum(x['area'] for x in rows if x['part'] in PIPE), 2)
    a_tie = round(sum(x['area'] for x in rows if x['part'] in TIE | CONE), 2)
    a_valve = round(sum(x['area'] for x in rows if x['part'] in VALVE), 2)
    l_elbow = round(sum(x['summ'] for x in rows if x['part'] in ELBOW), 2)
    l_pipe = straight_m
    l_tie = round(sum(x['summ'] for x in rows if x['part'] in TIE | CONE), 2)
    l_valve = round(sum(x['summ'] for x in rows if x['part'] in VALVE), 2)

    # основной типоразмер: по наибольшей длине среди Pipe (иначе среди всех)
    bysize = collections.defaultdict(float)
    devs = {}
    for x in rows:
        if x['od'] and x['id'] and x['dev']:
            key = (x['dn'], x['id'], x['ins'], x['od'])
            bysize[key] += x['work'] if x['part'] not in PIPE else x['summ']
            devs[key] = round(x['dev'])
    sizes = []
    for key, ln in sorted(bysize.items(), key=lambda kv: -kv[1]):
        dn, idd, ins, od = key
        sizes.append({'dn': dn, 'pipe_od': idd, 'ins': ins, 'clad_od': od,
                      'dev': devs[key], 'len': round(ln, 2)})
    main = sizes[0] if sizes else {}

    # короткое имя: 50150-HMP-150
    short = '-'.join(name.split('-')[:3])
    parts = name.split('-')
    medium = parts[1] if len(parts) > 1 else ''

    # ---------- ветки (branch) — по фактическому типоразмеру трубы ----------
    branches = {}
    main_key = None
    if sizes:
        m0 = sizes[0]
        main_key = f"{int(m0['pipe_od'])}-{int(m0['ins'])}-{int(m0['clad_od'])}"
    for x in rows:
        if x['od'] and x['id'] and x['dev']:
            key = f"{int(x['id'])}-{int(x['ins'])}-{int(x['od'])}"
        else:
            key = main_key
        if not key:
            continue
        b = branches.setdefault(key, {
            'id': key,
            'label': key.replace('-', '/'),
            'dn': dn_of(x['id'], x['dn']) if (x['od'] and x['id']) else (main['dn'] if main else None),
            'pipe_od': x['id'] if x['od'] else (main.get('pipe_od') if main else 0),
            'ins': x['ins'] if x['od'] else (main.get('ins') if main else 0),
            'clad_od': x['od'] if x['od'] else (main.get('clad_od') if main else 0),
            'dev': round(x['dev']) if x['dev'] else (main.get('dev') if main else 0),
            'al': set(),
            'length_m': 0.0, 'area_m2': 0.0, 'straight_m': 0.0,
            'elbows': 0, 'ties': 0, 'cones': 0, 'valves': 0, 'flanges': 0,
            'work': 0.0,
        })
        b['length_m'] += x['summ']
        b['area_m2'] += x['area']
        b['work'] += x['work']
        if x['al']:
            b['al'].add(re.sub(r'\s*мм', '', x['al']).strip())
        if x['part'] in PIPE:
            b['straight_m'] += x['summ']
        elif x['part'] in ELBOW:
            b['elbows'] += x['qty']
        elif x['part'] in TIE:
            b['ties'] += x['qty']
        elif x['part'] in CONE:
            b['cones'] += x['qty']
        elif x['part'] in VALVE:
            b['valves'] += x['qty']
        elif x['part'] in FLANGE:
            b['flanges'] += x['qty']

    blist = []
    for b in branches.values():
        b['al'] = sorted(b['al'])
        for k in ('length_m', 'area_m2', 'straight_m', 'work'):
            b[k] = round(b[k], 2)
        blist.append(b)
    blist.sort(key=lambda b: (-b['work'], -b['length_m']))

    # чемоданы раскладываем по веткам: по диаметру арматуры, остальное — на главную ветку
    if blist:
        for b in blist:
            b['cases'] = 0
        for bx in boxes_by_line.get(num, []):
            want = re.split(r'[/\\]', bx.get('dn') or '')[0].strip()
            tgt = None
            if want.isdigit():
                w = int(want)
                for b in blist:
                    if b.get('dn') and int(b['dn']) == w:
                        tgt = b
                        break
                if tgt is None:
                    for b in blist:
                        if b.get('pipe_od') and abs(int(b['pipe_od']) - w) <= 6:
                            tgt = b
                            break
            dest = tgt or blist[0]
            dest['cases'] += bx['qty']
            bx['branch'] = dest['id']
        # вес ветки не должен быть нулевым, иначе её отметки не влияют на процент
        for b in blist:
            if b['work'] <= 0:
                dv = (b['dev'] or 0) / 1000.0
                b['work'] = round(max(0.05, dv * (b['straight_m'] + 0.35 * b['elbows']
                                                  + 0.3 * (b['ties'] + b['cones'])
                                                  + 0.8 * b['cases'])), 2)

    dw = draw.get(num)

    out.append({
        'id': num,
        'drawing': (dw or {}).get('file'),
        'drawing_rev': (dw or {}).get('rev'),
        'name': name,
        'short': short,
        'num': num,
        'medium': medium,
        'oper_temp': L['otemp'],
        'design_temp': L['dtemp'],
        'partial': L['partial'],
        'release_date': (sorted(L['rel_dates'])[0] if L['rel_dates'] else None),
        'release_note': (' · '.join(sorted(L['rel_notes'])) if L['rel_notes'] else None),
        'release_pct': (round(sum(L['rel_pct']) / len(L['rel_pct']), 1) if L['rel_pct'] else None),
        'al': sorted(L['al']),
        'main': main,
        'sizes': sizes,
        'length_m': length_m,
        'area_m2': area_m2,
        'straight_m': straight_m,
        'elbows': elbows,
        'ties': ties,
        'cones': cones,
        'cases': cases,
        'valves_reg': valves_reg,
        'valves_kot': kot.get(num, 0),
        'boxes': boxes_by_line.get(num, []),
        'flanges': flanges,
        'connect': sorted(L['connect']),
        'branches': blist,
        'groups': {
            'elbow': {'area': a_elbow, 'len': l_elbow, 'qty': elbows,
                      'work': round(sum(x['work'] for x in rows if x['part'] in ELBOW), 2)},
            'pipe': {'area': a_pipe, 'len': l_pipe, 'qty': 0,
                     'work': round(sum(x['work'] for x in rows if x['part'] in PIPE), 2)},
            'fin': {'area': a_tie, 'len': l_tie, 'qty': ties + cones,
                    'work': round(sum(x['work'] for x in rows if x['part'] in TIE | CONE), 2)},
            'case': {'area': a_valve, 'len': l_valve, 'qty': cases,
                     'work': round(sum(x['work'] for x in rows if x['part'] in VALVE | FLANGE), 2)},
        },
        'rows': rows,
    })

# добор веса для групп без площади (нет размеров в строках VALVE и т.п.)
for l in out:
    dev = (l['main'].get('dev') or 0) / 1000.0
    g = l['groups']
    if g['case']['work'] == 0 and l['cases'] > 0 and dev:
        g['case']['work'] = round(0.8 * dev * l['cases'], 2)
    if g['elbow']['work'] == 0 and l['elbows'] > 0 and dev:
        g['elbow']['work'] = round(0.35 * dev * l['elbows'], 2)
    if g['fin']['work'] == 0 and (l['ties'] + l['cones']) > 0 and dev:
        g['fin']['work'] = round(0.3 * dev * (l['ties'] + l['cones']), 2)
    if g['pipe']['work'] == 0 and l['straight_m'] > 0 and dev:
        g['pipe']['work'] = round(dev * l['straight_m'], 2)

# связи: Connect содержит хвосты вида "154", "007" — разворачиваем в номера линий 50xxx
byname = {l['name']: l for l in out}
bynum = {l['num']: l for l in out}
for l in out:
    res = []
    for c in l['connect']:
        c2 = c.strip()
        if re.fullmatch(r'\d{3}', c2):
            cand = '50' + c2
            if cand in bynum:
                res.append(bynum[cand]['name'])
                continue
        if c2 in byname:
            res.append(byname[c2]['name'])
        elif re.fullmatch(r'50\d{3}', c2) and c2 in bynum:
            res.append(bynum[c2]['name'])
        else:
            res.append(c2)
    l['connect'] = sorted(set(x for x in res if x != l['name']))

# симметризация связей
cmap = collections.defaultdict(set)
for l in out:
    for c in l['connect']:
        if c in byname:
            cmap[l['name']].add(c)
            cmap[c].add(l['name'])
        else:
            cmap[l['name']].add(c)
for l in out:
    l['connect'] = sorted(cmap[l['name']])

meta = {
    'project': 'Valmet Tervasaari PM5 — изоляция трубопроводов',
    'source': 'TESA REGISTER - полный v65.xlsx + ЧЕМОДАНЫ.xlsx',
    'lines': len(out),
    'total_length_m': round(sum(l['length_m'] for l in out), 1),
    'total_area_m2': round(sum(l['area_m2'] for l in out), 1),
    'total_elbows': sum(l['elbows'] for l in out),
    'total_ties': sum(l['ties'] for l in out),
    'total_cones': sum(l['cones'] for l in out),
    'total_cases': sum(l['cases'] for l in out),
    'drawings': sum(1 for l in out if l.get('drawing')),
}
os.makedirs(os.path.dirname(OUT), exist_ok=True)
json.dump({'meta': meta, 'lines': out}, open(OUT, 'w', encoding='utf-8'),
          ensure_ascii=False, indent=1)
print(json.dumps(meta, ensure_ascii=False, indent=1))
print('чемоданы: линий', len([k for k in kot if k in bynum]), '· всего', sum(kot.values()))
